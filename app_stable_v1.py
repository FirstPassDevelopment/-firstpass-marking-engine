import streamlit as st
from docx import Document
import pandas as pd
import io
from openai import OpenAI
from dotenv import dotenv_values
from openpyxl.styles import Font, Alignment, PatternFill

# -----------------------
# ENV
# -----------------------

config = dotenv_values(".env")
api_key = config.get("OPENAI_API_KEY")

if not api_key:
    st.error("Missing API key")
    st.stop()

client = OpenAI(api_key=api_key.strip())

# -----------------------
# INIT
# -----------------------

st.set_page_config(page_title="FirstPass Marking Engine – Stable", layout="wide")
st.title("FirstPass Marking Engine – Stable")
st.caption("Version 1.0 – Stable")

if "results" not in st.session_state:
    st.session_state.results = None

if "counts" not in st.session_state:
    st.session_state.counts = None

# -----------------------
# ACARA
# -----------------------

ACARA_EXPECTATIONS = {
    "Year 7": "Students explain how language features create meaning.",
    "Year 8": "Students explain how language features shape meaning.",
    "Year 9": "Students analyse how language features influence meaning.",
    "Year 10": "Students evaluate how language features shape meaning."
}

# -----------------------
# TIGHT CALIBRATION EXEMPLARS
# -----------------------

EXEMPLARS = """
C+ (At Standard):
- Clear and understandable ideas
- Basic explanation of language features
- Some structure
- Limited depth

IMPORTANT: Most students fall here.

---

B (Above Standard):
- Clearly more developed than typical work
- Consistent explanation of HOW meaning is created
- Deliberate use of evidence
- Controlled structure

Only award B if it is CLEARLY above standard.

---

A (Exceptional):
- Rare
- Insightful, precise, highly controlled
- Sophisticated reasoning sustained throughout

If unsure, it is NOT an A.
"""

# -----------------------
# HELPERS
# -----------------------

def extract_text(file):
    doc = Document(file)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

# -----------------------
# PASS 1
# -----------------------

def summarise(text, task_type):

    prompt = f"""
Task Type: {task_type}

Summarise this response for marking.

Focus on:
- ideas
- language use
- structure
- depth

3–4 lines only.

{text}
"""

    res = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2
    )

    return res.choices[0].message.content

# -----------------------
# PASS 2 (FINAL CALIBRATION)
# -----------------------

def grade_with_context(summary, cohort_summaries, year, task_type):

    prompt = f"""
You are marking a {year} English class.

Task Type: {task_type}

Curriculum expectation:
{ACARA_EXPECTATIONS[year]}

REFERENCE STANDARD:
{EXEMPLARS}

---

CRITICAL GRADING RULE:

Start at C+.

Ask:
"Is this CLEARLY better than typical student work?"

If NO → keep C+
If YES → consider B
If EXTREMELY strong → consider A

---

DO NOT:
- reward fluency alone
- reward effort
- reward "nice writing"

FOCUS ON:
- depth of explanation
- clarity of reasoning
- control of ideas

---

Cohort:
{cohort_summaries}

---

Student:
{summary}

---

Output EXACTLY:

Overall Grade: [A+/A/A-/B+/B/B-/C+/C/C-/D/E]

Summary: [1 sentence]

Justification:
[2–3 sentences]

Strengths:
- point
- point

Areas for Improvement:
- point
- point
"""

    res = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.5
    )

    return res.choices[0].message.content

# -----------------------
# INSIGHTS
# -----------------------

def generate_cohort_insights(results):

    grades = [r["grade"] for r in results]

    c = sum(g.startswith("C") for g in grades)
    b = sum(g.startswith("B") for g in grades)

    dominant = "C" if c >= b else "B"

    return f"""
Grade Distribution Insight:
- Dominant band: {dominant}

Interpretation:
- {"Cohort aligned to expected standard" if dominant == "C" else "Slight upward drift into B band"}

Instructional Priorities:
1. Push explanation deeper
2. Strengthen use of evidence
3. Improve analytical clarity
"""

# -----------------------
# UI
# -----------------------

year = st.selectbox("Year", ["Year 7", "Year 8", "Year 9", "Year 10"])

task_type = st.selectbox("Task Type", [
    "Analytical",
    "Creative",
    "Hybrid",
    "Persuasive",
    "Short Response"
])

task_file = st.file_uploader("Task Sheet (.docx)", type=["docx"])
rubric_file = st.file_uploader("Rubric (.docx)", type=["docx"])
student_files = st.file_uploader("Student Files", type=["docx"], accept_multiple_files=True)

# -----------------------
# RUN
# -----------------------

status = st.empty()

if st.button("Run Marking"):

    status.info("Reading responses...")

    summaries = []

    for f in student_files:
        text = extract_text(f)
        s = summarise(text, task_type)
        summaries.append((f.name, s))

    status.info("Understanding cohort...")

    all_summaries = "\n\n".join([s[1] for s in summaries])

    results = []
    counts = {}

    status.info("Assigning grades...")

    for name, s in summaries:

        feedback = grade_with_context(s, all_summaries, year, task_type)

        grade = "C+"
        for g in ["A+","A","A-","B+","B","B-","C+","C","C-","D","E"]:
            if f"Overall Grade: {g}" in feedback:
                grade = g

        counts[grade] = counts.get(grade, 0) + 1

        results.append({
            "file": name,
            "grade": grade,
            "feedback": feedback
        })

    status.success("Complete")

    st.session_state.results = results
    st.session_state.counts = counts

# -----------------------
# DISPLAY
# -----------------------

if st.session_state.results:

    st.write("### Distribution")
    st.write(st.session_state.counts)

    st.write("### Cohort Insights")
    st.write(generate_cohort_insights(st.session_state.results))

    for i, r in enumerate(st.session_state.results):
        st.write(f"Student {i+1} | {r['grade']}")
        with st.expander("Feedback"):
            st.text(r["feedback"])

# -----------------------
# EXCEL EXPORT
# -----------------------

if st.session_state.results:

    df = pd.DataFrame(st.session_state.results)

    output = io.BytesIO()

    colour_map = {
        "A+": "00B050","A": "92D050","A-": "C6EFCE",
        "B+": "9BC2E6","B": "BDD7EE","B-": "DDEBF7",
        "C+": "FFF2CC","C": "FFE699","C-": "FFD966",
        "D": "F4B084","E": "FF9999"
    }

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name="Results")
        sheet = writer.sheets["Results"]

        for row in sheet.iter_rows(min_row=2):

            grade_val = row[1].value

            fill = PatternFill(
                start_color=colour_map.get(grade_val, "FFFFFF"),
                end_color=colour_map.get(grade_val, "FFFFFF"),
                fill_type="solid"
            )

            for i, cell in enumerate(row):

                if i == 0:
                    cell.font = Font(size=12)

                elif i == 1:
                    cell.font = Font(size=16, bold=True)
                    cell.alignment = Alignment(horizontal="center")

                else:
                    cell.font = Font(size=11)

                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 90

    st.download_button("Download Excel", output.getvalue(), "firstpass_results.xlsx")